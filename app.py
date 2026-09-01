import html
import math
import tempfile
from dataclasses import asdict, is_dataclass
from enum import Enum
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from item_analyzer import ItemAnalyzer


# ============================================================
# SAMPLE TEMPLATE FILES
# ============================================================
# These are the exact official template workbooks, checked into
# the repo under templates/. They are served as-is (not generated)
# so what the user downloads always matches what the analyzer's
# header-detection logic expects.

TEMPLATES_DIR = Path(__file__).parent / "templates"
ANSWER_KEY_TEMPLATE_PATH = TEMPLATES_DIR / "answer_key_template.xlsx"
STUDENT_TEMPLATE_PATH = TEMPLATES_DIR / "student_response_template.xlsx"


@st.cache_data
def load_template_bytes(path_str):

    path = Path(path_str)

    if not path.exists():
        return None

    return path.read_bytes()


# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Item Analysis",
    page_icon="📘",
    layout="wide",
    initial_sidebar_state="collapsed"
)


# ============================================================
# CUSTOM CSS
# ============================================================

st.markdown(
    """
<style>

@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500;600&display=swap');

html, body, [data-testid="stAppViewContainer"] {
    background-color: #F3F5F8 !important;
    color: #16233E !important;
    font-family: 'Inter', sans-serif !important;
}

h1, h2, h3 {
    font-family: 'Space Grotesk', sans-serif !important;
}

.header-box {
    background-color: #16233E;
    background-image:
        linear-gradient(
            rgba(255,255,255,0.04) 1px,
            transparent 1px
        ),
        linear-gradient(
            90deg,
            rgba(255,255,255,0.04) 1px,
            transparent 1px
        );

    background-size: 20px 20px;

    padding: 30px;

    border-radius: 14px;

    color: #FFFFFF;

    margin-bottom: 25px;
}

.badge-base {
    display: inline-block;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
}

.badge-retain {
    background-color: #E3F3EA;
    color: #2F8F5B;
}

.badge-review {
    background-color: #FBF0DE;
    color: #C9862B;
}

.badge-revise {
    background-color: #FAEAE2;
    color: #BD5B34;
}

.badge-discard {
    background-color: #FAE7E5;
    color: #B23A32;
}

.dist-container {
    display: flex;
    width: 100%;
    height: 30px;
    border-radius: 6px;
    overflow: hidden;
    margin: 10px 0;
    border: 1px solid #DEE4EF;
}

.dist-segment {
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-family: monospace;
    font-size: 11px;
    font-weight: bold;
}

.report-title {
    font-size: 18px;
    font-weight: 700;
    color: #16233E;
}

.small-muted {
    font-size: 12px;
    color: #65728A;
}

.template-card {
    background-color: #FFFFFF;
    border: 1px solid #DEE4EF;
    border-radius: 12px;
    padding: 18px 20px;
    margin-bottom: 10px;
}

.template-card-title {
    font-size: 14px;
    font-weight: 700;
    color: #16233E;
    margin-bottom: 6px;
}

.template-card-body {
    font-size: 13px;
    color: #4B5875;
    line-height: 1.5;
}

.template-required-badge {
    display: inline-block;
    background-color: #E6F1FB;
    color: #185FA5;
    font-size: 10px;
    font-weight: 700;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 20px;
    margin-left: 8px;
}

.template-error-card {
    background-color: #FAE7E5;
    border: 1px solid #F0C4BE;
    border-radius: 12px;
    padding: 16px 20px;
    margin: 10px 0;
}

.template-error-title {
    font-size: 14px;
    font-weight: 700;
    color: #B23A32;
    margin-bottom: 4px;
}

.template-error-body {
    font-size: 13px;
    color: #7A2A22;
    line-height: 1.5;
    white-space: pre-line;
}

.metric-info-card {
    background-color: #FFFFFF;
    border: 1px solid #E4E9F2;
    border-top: 4px solid var(--accent-color, #185FA5);
    border-radius: 14px;
    padding: 22px 22px 20px 22px;
    height: 100%;
    box-shadow: 0 2px 10px rgba(22, 35, 62, 0.04);
    transition: box-shadow 0.15s ease;
}

.metric-info-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 14px;
}

.metric-info-icon {
    display: flex;
    align-items: center;
    justify-content: center;
    width: 34px;
    height: 34px;
    min-width: 34px;
    border-radius: 10px;
    font-size: 16px;
    background-color: var(--accent-bg, #E6F1FB);
}

.metric-info-title {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 15.5px;
    font-weight: 700;
    color: #16233E;
    letter-spacing: -0.01em;
}

.metric-info-formula {
    font-family: 'JetBrains Mono', monospace;
    font-size: 12.5px;
    font-weight: 500;
    background-color: #F6F8FC;
    border: 1px solid #E4E9F2;
    border-left: 3px solid var(--accent-color, #185FA5);
    border-radius: 8px;
    padding: 12px 14px;
    color: #1E2D4E;
    margin-bottom: 14px;
    line-height: 1.6;
    white-space: pre-wrap;
}

.metric-info-body {
    font-family: 'Inter', sans-serif;
    font-size: 13.5px;
    font-weight: 400;
    color: #4B5875;
    line-height: 1.65;
    margin-bottom: 14px;
}

.metric-info-thresholds {
    font-family: 'Inter', sans-serif;
    font-size: 12.5px;
    color: #4B5875;
    display: flex;
    flex-direction: column;
    gap: 6px;
    padding-top: 12px;
    border-top: 1px dashed #E4E9F2;
}

.metric-info-thresholds .row {
    display: flex;
    align-items: center;
    gap: 8px;
}

.metric-info-thresholds code {
    font-family: 'JetBrains Mono', monospace;
    font-size: 11.5px;
    font-weight: 600;
    background-color: var(--accent-bg, #EEF2F9);
    color: var(--accent-color, #185FA5);
    padding: 3px 8px;
    border-radius: 20px;
    white-space: nowrap;
}

/* Force the three metric cards (and every nested Streamlit
   wrapper div in between) to stretch to one shared height, with
   the body copy growing to push each card's threshold list to
   the same bottom line regardless of text length. Descendant
   selectors (not direct-child) are used deliberately, since
   Streamlit nests several unnamed wrapper divs between the
   column and the actual markdown content. */

div[data-testid="stHorizontalBlock"]:has(.metric-info-card) {
    align-items: stretch;
}

div[data-testid="stHorizontalBlock"]:has(.metric-info-card)
    [data-testid="column"] {
    display: flex;
    height: 100%;
}

div[data-testid="stHorizontalBlock"]:has(.metric-info-card)
    [data-testid="stVerticalBlockBorderWrapper"],
div[data-testid="stHorizontalBlock"]:has(.metric-info-card)
    [data-testid="stVerticalBlock"],
div[data-testid="stHorizontalBlock"]:has(.metric-info-card)
    [data-testid="element-container"],
div[data-testid="stHorizontalBlock"]:has(.metric-info-card)
    [data-testid="stMarkdown"],
div[data-testid="stHorizontalBlock"]:has(.metric-info-card)
    [data-testid="stMarkdownContainer"] {
    width: 100%;
    height: 100%;
    display: flex;
    flex-direction: column;
    flex: 1 1 auto;
}

.metric-info-card {
    display: flex;
    flex-direction: column;
    flex: 1 1 auto;
    height: 100%;
}

.metric-info-body {
    flex: 1 1 auto;
}

/* Application shell and educational dashboard polish */
[data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 8% -8%, rgba(50, 118, 130, 0.10), transparent 29rem),
        radial-gradient(circle at 98% 8%, rgba(218, 163, 73, 0.10), transparent 24rem),
        #F7F8F5 !important;
}

[data-testid="stHeader"] {
    background: rgba(247, 248, 245, 0.82);
    backdrop-filter: blur(12px);
}

.block-container {
    max-width: 1240px;
    padding-top: 2.5rem;
    padding-bottom: 4rem;
}

h1, h2, h3 {
    color: #173B3F !important;
    letter-spacing: -0.03em;
}

.header-box {
    position: relative;
    overflow: hidden;
    background: linear-gradient(120deg, #173B3F, #1A5353);
    border: 1px solid rgba(255,255,255,0.13);
    border-radius: 22px;
    padding: 42px 46px;
    box-shadow: 0 18px 42px rgba(20, 54, 57, 0.18);
    margin-bottom: 30px;
}

.header-box::after {
    content: "";
    position: absolute;
    width: 340px;
    height: 340px;
    border: 1px solid rgba(248, 209, 130, 0.35);
    border-radius: 50%;
    right: -110px;
    top: -155px;
    box-shadow: 0 0 0 42px rgba(248, 209, 130, 0.06),
                0 0 0 84px rgba(248, 209, 130, 0.04);
}

.header-content { position: relative; z-index: 1; max-width: 720px; }

.eyebrow {
    color: #F8D182;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.13em;
    text-transform: uppercase;
}

.hero-title {
    color: #FFFFFF !important;
    font-size: clamp(2rem, 4vw, 3.05rem);
    line-height: 1.08;
    margin: 12px 0 13px;
}

.hero-copy { color: #D7E7E2; font-size: 15px; line-height: 1.7; margin: 0; }

.section-intro { margin: 34px 0 16px; }
.section-kicker {
    color: #A4641A;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-bottom: 5px;
}
.section-title {
    color: #173B3F;
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.55rem;
    font-weight: 700;
    letter-spacing: -0.025em;
    margin: 0;
}
.section-copy { color: #617477; font-size: 13.5px; line-height: 1.55; margin: 5px 0 0; }

.workflow-steps {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 10px;
    margin: 18px 0 16px;
}
.workflow-step {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 12px 14px;
    background: #FFFFFF;
    border: 1px solid #DCE7E2;
    border-radius: 12px;
    color: #52666A;
    font-size: 12.5px;
}
.workflow-step span {
    display: grid;
    place-items: center;
    width: 23px;
    height: 23px;
    border-radius: 50%;
    background: #E5F1EC;
    color: #176B5A;
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px;
    font-weight: 700;
}
.upload-panel-title {
    color: #173B3F;
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 700;
    font-size: 15px;
    margin-bottom: 7px;
}

[data-testid="stButton"] > button,
[data-testid="stDownloadButton"] > button {
    min-height: 2.75rem;
    border-radius: 10px;
    font-weight: 600;
    transition: transform 0.15s ease, box-shadow 0.15s ease;
}
[data-testid="stButton"] > button[kind="primary"],
[data-testid="stDownloadButton"] > button[kind="primary"] {
    background: #176B5A;
    border-color: #176B5A;
}
[data-testid="stButton"] > button:hover,
[data-testid="stDownloadButton"] > button:hover {
    transform: translateY(-1px);
    box-shadow: 0 7px 14px rgba(23, 107, 90, 0.16);
}
[data-testid="stMetric"] {
    background: #FFFFFF;
    border: 1px solid #DCE7E2;
    border-radius: 14px;
    padding: 16px 18px;
    box-shadow: 0 2px 8px rgba(23, 59, 63, 0.035);
}
[data-testid="stMetricLabel"] { color: #617477; font-size: 12px; font-weight: 600; }
[data-testid="stMetricValue"] { color: #173B3F; font-family: 'Space Grotesk', sans-serif; }
[data-testid="stExpander"] {
    background: #FFFFFF;
    border: 1px solid #DCE7E2;
    border-radius: 13px;
    overflow: hidden;
    margin-bottom: 9px;
}
[data-testid="stExpander"] summary { padding: 5px 4px; font-weight: 650; color: #173B3F; }
[data-testid="stDataFrame"] { border: 1px solid #DCE7E2; border-radius: 12px; overflow: hidden; }

/* Equal-height methodology cards and textbook-style formula notation. */
.metric-card-grid {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 22px;
    align-items: stretch;
}
.metric-card-grid .metric-info-card {
    min-height: 510px;
}
.formula-label {
    color: #173B3F;
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px;
    font-weight: 700;
    margin-bottom: 8px;
}
.formula-expression {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    color: #173B3F;
    font-family: 'JetBrains Mono', monospace;
    font-size: 11.5px;
    font-weight: 600;
    line-height: 1.35;
    text-align: center;
}
.formula-fraction {
    display: inline-grid;
    grid-template-rows: auto auto;
    min-width: 130px;
    vertical-align: middle;
}
.formula-fraction > span:first-child {
    border-bottom: 1.5px solid currentColor;
    padding: 0 5px 4px;
}
.formula-fraction > span:last-child { padding: 4px 5px 0; }
.formula-operator { font-size: 18px; line-height: 1; }
.download-callout {
    background: linear-gradient(115deg, #E7F3EE, #F9FCF9);
    border: 1px solid #C6E0D4;
    border-left: 5px solid #176B5A;
    border-radius: 14px;
    padding: 18px 20px;
    margin: 16px 0 10px;
}
.download-callout strong {
    display: block;
    color: #173B3F;
    font-family: 'Space Grotesk', sans-serif;
    font-size: 16px;
    margin-bottom: 4px;
}
.download-callout span { color: #52666A; font-size: 13px; line-height: 1.5; }

@media (max-width: 700px) {
    .block-container { padding-top: 1.2rem; }
    .header-box { padding: 30px 25px; border-radius: 17px; }
    .workflow-steps { grid-template-columns: 1fr; }
    .metric-card-grid { grid-template-columns: 1fr; }
    .metric-card-grid .metric-info-card { min-height: 0; }
}

</style>
""",
    unsafe_allow_html=True
)


# ============================================================
# HEADER
# ============================================================

st.html(
    """
<div class="header-box">
    <div class="header-content">
        <div class="eyebrow">Item Analysis</div>
        <h1 class="hero-title">Make every question a better teaching decision.</h1>
        <p class="hero-copy">
            Upload an answer key and student responses to turn classroom
            results into practical evidence: which items to retain, review,
            revise, or remove.
        </p>
    </div>
</div>
"""
)


# ============================================================
# METRIC METHODOLOGY (how the numbers are computed)
# ============================================================
#
# Shown up front, before the upload step, so users know exactly
# what the Difficulty Index, Discrimination Index, and Distractor
# Efficiency numbers mean before they see them on the dashboard.
# These formulas mirror item_analyzer.py exactly.

st.markdown(
    """
<div class="section-intro">
    <div class="section-kicker">Before you begin</div>
    <div class="section-title">Understand the evidence behind each item</div>
    <p class="section-copy">Three measures help you distinguish questions
    that are working well from those that need attention.</p>
</div>
""",
    unsafe_allow_html=True
)

st.markdown(
    """
<div class="metric-card-grid">
  <div class="metric-info-card" style="--accent-color:#185FA5; --accent-bg:#E6F1FB;">
    <div class="metric-info-header"><div class="metric-info-icon">🎯</div><div class="metric-info-title">Difficulty Index</div></div>
    <div class="metric-info-formula">
      <div class="formula-label">Difficulty</div>
      <div class="formula-expression"><div class="formula-fraction"><span>Correct responses</span><span>Total students</span></div></div>
    </div>
    <div class="metric-info-body">The share of students who answered the item correctly, from 0 to 1. Higher values indicate an easier item; lower values indicate a harder item.</div>
    <div class="metric-info-thresholds">
      <div class="row"><code>&lt; 0.20</code> Very difficult</div>
      <div class="row"><code>0.20 – 0.80</code> Ideal</div>
      <div class="row"><code>&gt; 0.80</code> Too easy</div>
    </div>
  </div>
  <div class="metric-info-card" style="--accent-color:#8A4FBE; --accent-bg:#F1E9FA;">
    <div class="metric-info-header"><div class="metric-info-icon">📈</div><div class="metric-info-title">Discrimination Index</div></div>
    <div class="metric-info-formula">
      <div class="formula-label">Top-group success − bottom-group success</div>
      <div class="formula-expression">
        <div class="formula-fraction"><span>Correct in top 27%</span><span>n</span></div>
        <span class="formula-operator">−</span>
        <div class="formula-fraction"><span>Correct in bottom 27%</span><span>n</span></div>
      </div>
    </div>
    <div class="metric-info-body">Compares the top-scoring and bottom-scoring 27% of students. It shows how well an item separates stronger students from weaker ones, on a scale of −1 to 1.</div>
    <div class="metric-info-thresholds">
      <div class="row"><code>&lt; 0</code> Negative</div>
      <div class="row"><code>0.00 – 0.19</code> Poor</div>
      <div class="row"><code>0.20 – 0.29</code> Fair</div>
      <div class="row"><code>≥ 0.30</code> Good</div>
    </div>
  </div>
  <div class="metric-info-card" style="--accent-color:#2F8F5B; --accent-bg:#E3F3EA;">
    <div class="metric-info-header"><div class="metric-info-icon">🧩</div><div class="metric-info-title">Distractor Efficiency</div></div>
    <div class="metric-info-formula">
      <div class="formula-label">Efficiency</div>
      <div class="formula-expression"><div class="formula-fraction"><span>Functional distractors</span><span>Total distractors</span></div><span class="formula-operator">× 100%</span></div>
    </div>
    <div class="metric-info-body">Each wrong option is checked for use. A distractor selected by at least 5% of students is functional; options selected by fewer than 5% are flagged for review.</div>
    <div class="metric-info-thresholds">
      <div class="row"><code>≥ 5%</code> selected → Functional</div>
      <div class="row"><code>&lt; 5%</code> selected → Non-functional (flagged)</div>
    </div>
  </div>
</div>
""",
    unsafe_allow_html=True
)

st.markdown("")


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def sanitize(data):
    """
    Convert dataclasses and Enum values into JSON/Streamlit-safe
    Python objects.

    Also handles NaN and Infinity.
    """

    # IMPORTANT:
    # This fixes the previous:
    # NameError: name 'Enum' is not defined
    if isinstance(data, Enum):
        return data.value

    if is_dataclass(data):

        return sanitize(
            asdict(data)
        )

    if isinstance(data, dict):

        return {
            key: sanitize(value)
            for key, value in data.items()
        }

    if isinstance(data, list):

        return [
            sanitize(item)
            for item in data
        ]

    if isinstance(data, tuple):

        return [
            sanitize(item)
            for item in data
        ]

    if isinstance(data, float):

        if math.isnan(data) or math.isinf(data):

            return 0.0

    return data


def get_value(
    obj,
    key,
    default=None
):

    if isinstance(obj, dict):

        return obj.get(
            key,
            default
        )

    return getattr(
        obj,
        key,
        default
    )


def recommendation_badge(
    recommendation
):

    rec = str(
        recommendation
    ).upper()

    if rec == "RETAIN":

        return (
            '<span class="badge-base badge-retain">'
            'RETAIN'
            '</span>'
        )

    if rec == "REVIEW":

        return (
            '<span class="badge-base badge-review">'
            'REVIEW'
            '</span>'
        )

    if rec == "REVISE":

        return (
            '<span class="badge-base badge-revise">'
            'REVISE'
            '</span>'
        )

    if rec == "DISCARD":

        return (
            '<span class="badge-base badge-discard">'
            'DISCARD'
            '</span>'
        )

    return (
        '<span class="badge-base badge-review">'
        f'{html.escape(rec)}'
        '</span>'
    )


# ============================================================
# EXCEL REPORT GENERATOR
# ============================================================

def create_excel_report(results):

    """
    Create a complete downloadable Excel report.

    Sheet order:

    1. Item Analysis
    2. Option Breakdown
    3. Overall Summary

    Each sheet is stamped with an "EXAM INFORMATION" banner
    (Examination Name / Year / Class / Subject) pulled from the
    student response file the user uploaded, followed by a blank
    spacer row and then the data table.
    """

    output = BytesIO()

    summary = get_value(
        results,
        "summary",
        {}
    )

    items = get_value(
        results,
        "items",
        []
    )

    exam_info = get_value(
        results,
        "exam_info",
        {}
    )

    exam_info_rows = [
        (
            "Examination Name :",
            get_value(exam_info, "Examination Name", "")
        ),
        (
            "Examination Year :",
            get_value(exam_info, "Examination Year", "")
        ),
        (
            "Class :",
            get_value(exam_info, "Class", "")
        ),
        (
            "Subject :",
            get_value(exam_info, "Subject", "")
        ),
    ]

    # Banner occupies rows 1-6 (EXAM INFORMATION, blank, then the
    # 4 fields), row 7 is a spacer, and the data table header
    # lands on row 8 — so pandas' 0-indexed startrow is 7.
    DATA_STARTROW = 7

    # --------------------------------------------------------
    # ITEM ANALYSIS DATA
    # --------------------------------------------------------

    item_rows = []

    for item in items:

        recommendation = get_value(
            item,
            "recommendation",
            ""
        )

        if isinstance(
            recommendation,
            Enum
        ):

            recommendation = (
                recommendation.value
            )

        nfd = get_value(
            item,
            "non_functional_distractors",
            []
        )

        if nfd is None:
            nfd = []

        item_rows.append({

            "Question":
                get_value(
                    item,
                    "question",
                    ""
                ),

            "Question Number":
                get_value(
                    item,
                    "question_number",
                    ""
                ),

            "Correct Answer":
                get_value(
                    item,
                    "correct_answer",
                    ""
                ),

            "Correct Count":
                get_value(
                    item,
                    "correct_count",
                    0
                ),

            "Total Students":
                get_value(
                    item,
                    "total_students",
                    0
                ),

            "Difficulty Index":
                round(
                    get_value(
                        item,
                        "difficulty",
                        0
                    ),
                    2
                ),

            "Difficulty Status":
                get_value(
                    item,
                    "difficulty_status",
                    ""
                ),

            "Discrimination Index":
                round(
                    get_value(
                        item,
                        "discrimination",
                        0
                    ),
                    2
                ),

            "Discrimination Status":
                get_value(
                    item,
                    "discrimination_status",
                    ""
                ),

            "Distractor Efficiency (%)":
                round(
                    get_value(
                        item,
                        "distractor_efficiency",
                        0
                    ),
                    2
                ),

            "Non-functional Distractors":
                ", ".join(
                    str(x)
                    for x in nfd
                )
                if nfd
                else "None",

            "Recommendation":
                str(
                    recommendation
                ).upper()
        })

    item_df = pd.DataFrame(
        item_rows
    )

    # --------------------------------------------------------
    # OPTION BREAKDOWN
    # --------------------------------------------------------

    option_rows = []

    for item in items:

        question = get_value(
            item,
            "question",
            ""
        )

        breakdown = get_value(
            item,
            "option_breakdown",
            []
        )

        for option in breakdown:

            is_correct = bool(
                get_value(
                    option,
                    "is_correct",
                    False
                )
            )

            is_functional = bool(
                get_value(
                    option,
                    "is_functional",
                    False
                )
            )

            if is_correct:

                status = "Correct Answer"

            elif is_functional:

                status = "Functional Distractor"

            else:

                status = "Non-functional Distractor"

            option_rows.append({

                "Question":
                    question,

                "Option":
                    get_value(
                        option,
                        "option",
                        ""
                    ),

                "Selection Count":
                    get_value(
                        option,
                        "count",
                        0
                    ),

                "Percentage":
                    round(
                        get_value(
                            option,
                            "percentage",
                            0
                        ),
                        2
                    ),

                "Status":
                    status
            })

    option_df = pd.DataFrame(
        option_rows
    )

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    summary_df = pd.DataFrame({

        "Measure": [

            "Total Students",

            "Total Evaluated Items",

            "Mean Score",

            "Standard Deviation",

            "Minimum Score",

            "Maximum Score",

            "Mean Percentage"
        ],

        "Value": [

            get_value(
                summary,
                "total_students",
                0
            ),

            get_value(
                summary,
                "total_questions",
                0
            ),

            round(
                get_value(
                    summary,
                    "mean_score",
                    0
                ),
                2
            ),

            round(
                get_value(
                    summary,
                    "std_score",
                    0
                ),
                2
            ),

            get_value(
                summary,
                "min_score",
                0
            ),

            get_value(
                summary,
                "max_score",
                0
            ),

            round(
                get_value(
                    summary,
                    "mean_percentage",
                    0
                ),
                2
            )
        ]
    })

    # --------------------------------------------------------
    # WRITE EXCEL
    # --------------------------------------------------------

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        item_df.to_excel(
            writer,
            sheet_name="Item Analysis",
            index=False,
            startrow=DATA_STARTROW
        )

        option_df.to_excel(
            writer,
            sheet_name="Option Breakdown",
            index=False,
            startrow=DATA_STARTROW
        )

        summary_df.to_excel(
            writer,
            sheet_name="Overall Summary",
            index=False,
            startrow=DATA_STARTROW
        )

        # ----------------------------------------------------
        # EXAM INFO BANNER + FORMATTING
        # ----------------------------------------------------

        workbook = writer.book

        for worksheet in workbook.worksheets:

            # Shared workbook palette: calm, readable, and aligned
            # with the education-focused application interface.
            navy_fill = PatternFill(
                "solid",
                fgColor="173B3F"
            )
            teal_fill = PatternFill(
                "solid",
                fgColor="176B5A"
            )
            soft_teal_fill = PatternFill(
                "solid",
                fgColor="E7F3EE"
            )
            stripe_fill = PatternFill(
                "solid",
                fgColor="F4F8F6"
            )
            label_fill = PatternFill(
                "solid",
                fgColor="D9EBE4"
            )
            gold_fill = PatternFill(
                "solid",
                fgColor="FFF1D6"
            )
            retain_fill = PatternFill(
                "solid",
                fgColor="D9F0E2"
            )
            review_fill = PatternFill(
                "solid",
                fgColor="FFF1D6"
            )
            revise_fill = PatternFill(
                "solid",
                fgColor="FCE8D5"
            )
            discard_fill = PatternFill(
                "solid",
                fgColor="F8DDDA"
            )
            thin_teal = Side(
                style="thin",
                color="BFD8CE"
            )
            header_row = DATA_STARTROW + 1
            first_data_row = header_row + 1
            last_column = worksheet.max_column

            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.tabColor = (
                "176B5A"
                if worksheet.title != "Overall Summary"
                else "D69E2E"
            )

            # ------------------------------------------------
            # EXAM INFORMATION BANNER (rows 1-6)
            # ------------------------------------------------

            title_cell = worksheet.cell(
                row=1,
                column=1,
                value="EXAM INFORMATION"
            )

            worksheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=last_column,
            )

            title_cell.fill = navy_fill
            title_cell.font = Font(
                bold=True,
                size=14,
                color="FFFFFF",
                name="Aptos Display"
            )
            title_cell.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )
            worksheet.row_dimensions[1].height = 28

            for row_offset, (label, value) in enumerate(
                exam_info_rows,
                start=3
            ):

                label_cell = worksheet.cell(
                    row=row_offset,
                    column=1,
                    value=label
                )

                label_cell.fill = label_fill
                label_cell.font = Font(
                    bold=True,
                    color="173B3F",
                    name="Aptos"
                )
                label_cell.alignment = Alignment(
                    vertical="center"
                )

                value_cell = worksheet.cell(
                    row=row_offset,
                    column=2,
                    value=value
                )
                value_cell.font = Font(
                    color="314E52",
                    name="Aptos"
                )
                value_cell.alignment = Alignment(
                    vertical="center"
                )
                worksheet.row_dimensions[row_offset].height = 20

            # Table headers establish a strong visual starting point.
            for header_cell in worksheet[header_row]:

                header_cell.fill = teal_fill
                header_cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                    name="Aptos"
                )
                header_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                header_cell.border = Border(
                    top=thin_teal,
                    bottom=thin_teal,
                )

            worksheet.row_dimensions[header_row].height = 32

            # Gentle zebra striping makes longer tables easier to scan.
            for row in range(
                first_data_row,
                worksheet.max_row + 1,
            ):

                for cell in worksheet[row]:

                    cell.font = Font(
                        color="243D41",
                        name="Aptos",
                        size=10,
                    )
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = Border(
                        bottom=Side(
                            style="hair",
                            color="D8E5DF",
                        )
                    )

                    if row % 2 == 1:
                        cell.fill = stripe_fill

            # Freeze panes just below the data table's header
            # row (banner rows 1-7 + header row 8).
            worksheet.freeze_panes = (
                f"A{DATA_STARTROW + 2}"
            )
            worksheet.auto_filter.ref = (
                f"A{header_row}:"
                f"{worksheet.cell(worksheet.max_row, last_column).coordinate}"
            )

            # Auto-size columns
            for column_index, column_cells in enumerate(
                worksheet.columns,
                start=1,
            ):

                max_length = 0

                column_letter = get_column_letter(
                    column_index
                )

                for cell in column_cells:

                    try:

                        cell_length = len(
                            str(
                                cell.value
                            )
                        )

                        max_length = max(
                            max_length,
                            cell_length
                        )

                    except Exception:

                        pass

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max(
                        max_length + 2,
                        12
                    ),
                    40
                )

            # Improve readability for the descriptive first column
            # and keep compact numeric fields centrally aligned.
            worksheet.column_dimensions["A"].width = min(
                max(
                    worksheet.column_dimensions["A"].width,
                    20,
                ),
                36,
            )

            for column in range(2, last_column + 1):

                for cell in worksheet.iter_cols(
                    min_col=column,
                    max_col=column,
                    min_row=first_data_row,
                    max_row=worksheet.max_row,
                ):

                    for value_cell in cell:

                        if isinstance(value_cell.value, (int, float)):
                            value_cell.alignment = Alignment(
                                horizontal="center",
                                vertical="center",
                            )

            # Display decimal-valued analysis results with two places,
            # including trailing zeroes (for example, 0.50).
            decimal_headers = {
                "Difficulty Index",
                "Discrimination Index",
                "Distractor Efficiency (%)",
                "Percentage",
            }

            for header_cell in worksheet[DATA_STARTROW + 1]:

                if header_cell.value in decimal_headers:

                    for column_cells in worksheet.iter_cols(
                        min_col=header_cell.column,
                        max_col=header_cell.column,
                        min_row=DATA_STARTROW + 2,
                        max_row=worksheet.max_row,
                    ):

                        for cell in column_cells:

                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "0.00"

            if worksheet.title == "Overall Summary":

                for row in range(
                    DATA_STARTROW + 2,
                    worksheet.max_row + 1,
                ):

                    if worksheet.cell(row=row, column=1).value in {
                        "Mean Score",
                        "Standard Deviation",
                        "Mean Percentage",
                    }:

                        worksheet.cell(
                            row=row,
                            column=2,
                        ).number_format = "0.00"

            # Use color cues only for meaningful diagnostic labels.
            for header_cell in worksheet[header_row]:

                if header_cell.value not in {
                    "Recommendation",
                    "Status",
                }:
                    continue

                for row in range(
                    first_data_row,
                    worksheet.max_row + 1,
                ):

                    status_cell = worksheet.cell(
                        row=row,
                        column=header_cell.column,
                    )
                    status = str(status_cell.value).upper()

                    if "RETAIN" in status or "CORRECT" in status:
                        status_cell.fill = retain_fill
                        status_cell.font = Font(
                            bold=True,
                            color="176B5A",
                            name="Aptos",
                        )

                    elif "NON-FUNCTIONAL" in status:
                        status_cell.fill = discard_fill
                        status_cell.font = Font(
                            bold=True,
                            color="B42318",
                            name="Aptos",
                        )

                    elif "REVIEW" in status or "FUNCTIONAL" in status:
                        status_cell.fill = review_fill
                        status_cell.font = Font(
                            bold=True,
                            color="9A6700",
                            name="Aptos",
                        )

                    elif "REVISE" in status:
                        status_cell.fill = revise_fill
                        status_cell.font = Font(
                            bold=True,
                            color="B54708",
                            name="Aptos",
                        )

                    elif "DISCARD" in status:
                        status_cell.fill = discard_fill
                        status_cell.font = Font(
                            bold=True,
                            color="B42318",
                            name="Aptos",
                        )

            if worksheet.title == "Overall Summary":

                for row in range(
                    first_data_row,
                    worksheet.max_row + 1,
                ):

                    worksheet.cell(
                        row=row,
                        column=1,
                    ).fill = soft_teal_fill

                    value_cell = worksheet.cell(
                        row=row,
                        column=2,
                    )
                    value_cell.fill = gold_fill
                    value_cell.font = Font(
                        bold=True,
                        color="173B3F",
                        name="Aptos",
                    )

    output.seek(0)

    return output.getvalue()


# ============================================================
# FILE UPLOAD (centered popup, template-guided)
# ============================================================
#
# Each slot shows a single upload button. Clicking it opens a
# centered modal popup (st.dialog) that strictly explains the
# required template, offers the sample download, and only lets
# the user proceed once a file has been chosen. Nothing about
# the format is buried inline in the page — it's front-and-center
# in the popup every time.

st.markdown(
    """
<div class="section-intro">
    <div class="section-kicker">Analysis workflow</div>
    <div class="section-title">Prepare your assessment files</div>
    <p class="section-copy">Use the supplied templates so every response is
    read accurately and your results remain comparable.</p>
</div>
<div class="workflow-steps">
    <div class="workflow-step"><span>1</span>Download the templates</div>
    <div class="workflow-step"><span>2</span>Upload both completed files</div>
    <div class="workflow-step"><span>3</span>Review and export insights</div>
</div>
""",
    unsafe_allow_html=True
)


def _init_upload_state(prefix):

    if f"{prefix}_confirmed_file" not in st.session_state:
        st.session_state[f"{prefix}_confirmed_file"] = None


def _render_template_popup_body(
    prefix,
    label,
    description_html,
    template_path,
    template_download_name
):
    """
    Shared content rendered inside the centered modal popup for
    a given upload slot (Answer Key / Student Responses).
    """

    # --------------------------------------------------------
    # STRICT TEMPLATE NOTICE
    # --------------------------------------------------------

    st.warning(
        f"⚠️ **Strict template required.** Your {label.lower()} "
        "file **must** match the format below exactly, or the "
        "analysis will fail. Please download and use the sample "
        "template rather than a file of your own formatting."
    )

    st.markdown(
        f"""
<div class="template-card">
    <div class="template-card-title">
        {label}
        <span class="template-required-badge">Template required</span>
    </div>
    <div class="template-card-body">
        {description_html}
    </div>
</div>
""",
        unsafe_allow_html=True
    )

    # --------------------------------------------------------
    # TEMPLATE DOWNLOAD
    # --------------------------------------------------------

    template_bytes = load_template_bytes(str(template_path))

    if template_bytes is not None:

        st.download_button(
            label=f"⬇️ Download sample {label} template",
            data=template_bytes,
            file_name=template_download_name,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            key=f"{prefix}_template_download",
            use_container_width=True
        )

    else:

        st.warning(
            f"Sample template file not found at {template_path}."
        )

    st.markdown("---")

    # --------------------------------------------------------
    # UPLOADER
    # --------------------------------------------------------

    pending_file = st.file_uploader(
        f"Upload your {label} spreadsheet (must match the template above)",
        type=["xlsx", "xls"],
        key=f"{prefix}_uploader_dialog"
    )

    btn_col, cancel_col = st.columns(2)

    with btn_col:

        submit_clicked = st.button(
            "Use this file",
            key=f"{prefix}_submit_btn",
            type="primary",
            disabled=pending_file is None,
            use_container_width=True
        )

    with cancel_col:

        cancel_clicked = st.button(
            "Cancel",
            key=f"{prefix}_cancel_btn",
            use_container_width=True
        )

    if submit_clicked and pending_file is not None:

        st.session_state[f"{prefix}_confirmed_file"] = pending_file
        st.rerun()

    if cancel_clicked:

        st.rerun()


@st.dialog("📋 Answer Key — Template Required", width="large")
def _open_answer_key_dialog():

    _render_template_popup_body(
        prefix="key",
        label="Answer Key",
        description_html=(
            "One row per question with a <code>Question</code> "
            "column and a <code>Correct Answer</code> column. "
            "Answers must be <code>A</code>, <code>B</code>, "
            "<code>C</code>, or <code>D</code>."
        ),
        template_path=ANSWER_KEY_TEMPLATE_PATH,
        template_download_name="Answer_Key_Template.xlsx"
    )


@st.dialog("📋 Student Responses — Template Required", width="large")
def _open_student_responses_dialog():

    _render_template_popup_body(
        prefix="data",
        label="Student Responses",
        description_html=(
            "One row per student with <code>Student ID</code> "
            "and <code>Name</code> columns, followed by one "
            "column per question, matching the Answer Key."
        ),
        template_path=STUDENT_TEMPLATE_PATH,
        template_download_name="Student_Responses_Template.xlsx"
    )


def render_upload_slot(prefix, label, dialog_opener):
    """
    Renders one of the two upload slots (Answer Key / Student
    Responses):

      1. Just a button, by default.
      2. Clicking it opens a centered modal popup with the
         template rules, download link, uploader, and a
         confirm/cancel choice.
      3. Once confirmed, collapses to a short "uploaded" summary
         with a "Change file" option that reopens the popup.
    """

    _init_upload_state(prefix)

    confirmed_file = st.session_state[f"{prefix}_confirmed_file"]

    # ------------------------------------------------------
    # STATE 1: FILE ALREADY CONFIRMED
    # ------------------------------------------------------

    if confirmed_file is not None:

        st.success(
            f"✅ {label}: **{confirmed_file.name}**"
        )

        if st.button(
            f"Change {label.lower()} file",
            key=f"{prefix}_change_btn",
            use_container_width=True
        ):

            st.session_state[f"{prefix}_confirmed_file"] = None
            dialog_opener()

        return confirmed_file

    # ------------------------------------------------------
    # STATE 2: SHOW ONLY THE UPLOAD BUTTON — OPENS POPUP
    # ------------------------------------------------------

    if st.button(
        f"📤 Upload {label}",
        key=f"{prefix}_open_btn",
        type="primary",
        use_container_width=True
    ):

        dialog_opener()

    return None


col_key, col_data = st.columns(2)

with col_key:

    st.markdown(
        '<div class="upload-panel-title">1. Answer key</div>',
        unsafe_allow_html=True
    )

    key_file = render_upload_slot(
        prefix="key",
        label="Answer Key",
        dialog_opener=_open_answer_key_dialog
    )

with col_data:

    st.markdown(
        '<div class="upload-panel-title">2. Student responses</div>',
        unsafe_allow_html=True
    )

    data_file = render_upload_slot(
        prefix="data",
        label="Student Responses",
        dialog_opener=_open_student_responses_dialog
    )


# ============================================================
# ANALYSIS
# ============================================================

if key_file and data_file:

    if st.button(
        "Run Analysis",
        type="primary",
        use_container_width=True
    ):

        with st.spinner(
            "Processing analysis files..."
        ):

            try:

                # ====================================================
                # TEMPORARY FILES
                # ====================================================

                with tempfile.TemporaryDirectory() as tmp_dir:

                    tmp_path = Path(
                        tmp_dir
                    )

                    key_path = (
                        tmp_path
                        / "key.xlsx"
                    )

                    data_path = (
                        tmp_path
                        / "data.xlsx"
                    )

                    key_path.write_bytes(
                        key_file.getvalue()
                    )

                    data_path.write_bytes(
                        data_file.getvalue()
                    )

                    # =================================================
                    # ANALYZER
                    # =================================================

                    analyzer = ItemAnalyzer()

                    raw_results = (
                        analyzer.run_analysis(
                            str(key_path),
                            str(data_path)
                        )
                    )

                # ====================================================
                # SANITIZE
                # ====================================================

                results = sanitize(
                    raw_results
                )

                # Store results in session state
                # so they remain available after download.
                st.session_state[
                    "analysis_results"
                ] = results

                st.success(
                    "✅ Analysis completed successfully."
                )

            except ValueError as e:

                st.markdown(
                    f"""
<div class="template-error-card">
    <div class="template-error-title">
        ⚠️ File doesn't match the required template
    </div>
    <div class="template-error-body">{html.escape(str(e))}</div>
</div>
""",
                    unsafe_allow_html=True
                )

                st.caption(
                    "Download the sample templates above, fill in "
                    "your data using the same columns, and "
                    "re-upload."
                )

            except Exception as e:

                st.error(
                    f"❌ Computational Execution Failure: {e}"
                )

                with st.expander(
                    "🔍 Technical Error Details",
                    expanded=True
                ):

                    st.exception(e)


# ============================================================
# DISPLAY RESULTS
# ============================================================

if (
    "analysis_results"
    in st.session_state
):

    results = st.session_state[
        "analysis_results"
    ]

    # ========================================================
    # GLOBAL SUMMARY
    # ========================================================

    summary = get_value(
        results,
        "summary",
        {}
    )

    st.markdown(
        """
<div class="section-intro">
    <div class="section-kicker">Assessment overview</div>
    <div class="section-title">Your cohort at a glance</div>
    <p class="section-copy">Start here for a quick picture of participation,
    score spread, and overall performance.</p>
</div>
""",
        unsafe_allow_html=True
    )

    metric_col1, metric_col2, metric_col3 = st.columns(3)

    with metric_col1:

        st.metric(
            "Total Students",
            get_value(
                summary,
                "total_students",
                0
            )
        )

    with metric_col2:

        st.metric(
            "Total Evaluated Items",
            get_value(
                summary,
                "total_questions",
                0
            )
        )

    with metric_col3:

        mean_score = get_value(
            summary,
            "mean_score",
            0
        )

        total_questions = get_value(
            summary,
            "total_questions",
            0
        )

        st.metric(
            "Mean Score",
            f"{float(mean_score):.2f} / {total_questions}"
        )

    # ========================================================
    # SECOND SUMMARY ROW
    # ========================================================

    metric_col4, metric_col5, metric_col6 = st.columns(3)

    with metric_col4:

        st.metric(
            "Standard Deviation",
            f"{float(get_value(summary, 'std_score', 0)):.2f}"
        )

    with metric_col5:

        st.metric(
            "Minimum Score",
            get_value(
                summary,
                "min_score",
                0
            )
        )

    with metric_col6:

        st.metric(
            "Maximum Score",
            get_value(
                summary,
                "max_score",
                0
            )
        )

    # ========================================================
    # ITEM ANALYSIS
    # ========================================================

    st.markdown("---")

    st.markdown(
        """
<div class="section-intro">
    <div class="section-kicker">Question review</div>
    <div class="section-title">Item-by-item analysis</div>
</div>
""",
        unsafe_allow_html=True
    )

    st.caption(
        "Each question is evaluated individually for difficulty, "
        "discrimination, distractor efficiency, and final recommendation."
    )

    items = get_value(
        results,
        "items",
        []
    )

    if not isinstance(
        items,
        list
    ):

        items = list(items)

    # ========================================================
    # RECOMMENDATION SUMMARY
    # ========================================================

    retain_count = 0
    review_count = 0
    revise_count = 0
    discard_count = 0

    for item in items:

        recommendation = str(
            get_value(
                item,
                "recommendation",
                ""
            )
        ).upper()

        if recommendation == "RETAIN":
            retain_count += 1

        elif recommendation == "REVIEW":
            review_count += 1

        elif recommendation == "REVISE":
            revise_count += 1

        elif recommendation == "DISCARD":
            discard_count += 1

    rec1, rec2, rec3, rec4 = st.columns(4)

    with rec1:
        st.metric(
            "✅ Retain",
            retain_count
        )

    with rec2:
        st.metric(
            "🟡 Review",
            review_count
        )

    with rec3:
        st.metric(
            "🟠 Revise",
            revise_count
        )

    with rec4:
        st.metric(
            "🔴 Discard",
            discard_count
        )

    # ========================================================
    # QUESTION LOOP
    # ========================================================

    for index, item in enumerate(
        items,
        start=1
    ):

        question = get_value(
            item,
            "question",
            f"Question {index}"
        )

        recommendation = get_value(
            item,
            "recommendation",
            "REVIEW"
        )

        recommendation = str(
            recommendation
        ).upper()

        # ----------------------------------------------------
        # EXPANDER
        # ----------------------------------------------------

        with st.expander(
            f"{question} — {recommendation}",
            expanded=False
        ):

            # =================================================
            # RECOMMENDATION
            # =================================================

            st.markdown(
                recommendation_badge(
                    recommendation
                ),
                unsafe_allow_html=True
            )

            st.markdown("")

            # =================================================
            # BASIC ITEM INFORMATION
            # =================================================

            info1, info2, info3 = st.columns(3)

            with info1:

                st.write(
                    f"**Correct Answer:** "
                    f"`{get_value(item, 'correct_answer', '')}`"
                )

            with info2:

                st.write(
                    f"**Correct Responses:** "
                    f"`{get_value(item, 'correct_count', 0)}` "
                    f"/ "
                    f"`{get_value(item, 'total_students', 0)}`"
                )

            with info3:

                st.write(
                    f"**Omitted Responses:** "
                    f"`{get_value(item, 'omitted_count', 0)}` "
                    f""
                )

            # =================================================
            # METRICS
            # =================================================

            st.markdown("---")

            gauge1, gauge2 = st.columns(2)

            # -------------------------------------------------
            # DIFFICULTY
            # -------------------------------------------------

            with gauge1:

                difficulty = get_value(
                    item,
                    "difficulty",
                    0.0
                )

                try:

                    difficulty = float(
                        difficulty
                    )

                except (
                    TypeError,
                    ValueError
                ):

                    difficulty = 0.0

                difficulty_status = get_value(
                    item,
                    "difficulty_status",
                    "N/A"
                )

                st.metric(
                    "Difficulty Index",
                    f"{difficulty:.2f}",
                    delta=str(
                        difficulty_status
                    ),
                    delta_color="off"
                )

                st.progress(
                    max(
                        0.0,
                        min(
                            1.0,
                            difficulty
                        )
                    )
                )

                st.caption(
                    "Proportion of students who answered correctly."
                )

            # -------------------------------------------------
            # DISCRIMINATION
            # -------------------------------------------------

            with gauge2:

                discrimination = get_value(
                    item,
                    "discrimination",
                    0.0
                )

                try:

                    discrimination = float(
                        discrimination
                    )

                except (
                    TypeError,
                    ValueError
                ):

                    discrimination = 0.0

                discrimination_status = get_value(
                    item,
                    "discrimination_status",
                    "N/A"
                )

                st.metric(
                    "Discrimination Index",
                    f"{discrimination:.2f}",
                    delta=str(
                        discrimination_status
                    ),
                    delta_color="off"
                )

                # Convert -1..+1 to 0..1
                normalized_discrimination = (
                    discrimination + 1
                ) / 2

                st.progress(
                    max(
                        0.0,
                        min(
                            1.0,
                            normalized_discrimination
                        )
                    )
                )

                st.caption(
                    "Difference between upper and lower group performance."
                )

            # =================================================
            # DISTRACTOR METRICS
            # =================================================

            st.markdown("---")

            extra1, extra2 = st.columns(2)

            with extra1:

                efficiency = get_value(
                    item,
                    "distractor_efficiency",
                    0.0
                )

                try:

                    efficiency = float(
                        efficiency
                    )

                except (
                    TypeError,
                    ValueError
                ):

                    efficiency = 0.0

                st.metric(
                    "Distractor Efficiency",
                    f"{efficiency:.0f}%"
                )

            with extra2:

                nfd_list = get_value(
                    item,
                    "non_functional_distractors",
                    []
                )

                if nfd_list:

                    st.write(
                        "**Non-functional Distractors:**"
                    )

                    st.error(
                        ", ".join(
                            str(x)
                            for x in nfd_list
                        )
                    )

                else:

                    st.write(
                        "**Non-functional Distractors:**"
                    )

                    st.success(
                        "None"
                    )

            # =================================================
            # RESPONSE DISTRIBUTION
            # =================================================

            st.markdown("---")

            st.markdown(
                "#### Response Distribution"
            )

            breakdown = get_value(
                item,
                "option_breakdown",
                []
            )

            if breakdown is None:

                breakdown = []

            # -------------------------------------------------
            # BAR
            # -------------------------------------------------

            bar_html = (
                '<div class="dist-container">'
            )

            for option in breakdown:

                percentage = get_value(
                    option,
                    "percentage",
                    0
                )

                option_label = get_value(
                    option,
                    "option",
                    ""
                )

                try:

                    percentage = float(
                        percentage
                    )

                except (
                    TypeError,
                    ValueError
                ):

                    percentage = 0.0

                if percentage <= 0:

                    continue

                is_correct = bool(
                    get_value(
                        option,
                        "is_correct",
                        False
                    )
                )

                is_functional = bool(
                    get_value(
                        option,
                        "is_functional",
                        False
                    )
                )

                if is_correct:

                    background = "#2F8F5B"

                elif is_functional:

                    background = "#2C5F8A"

                else:

                    background = "#B23A32"

                label = (
                    str(option_label)
                    if percentage >= 5
                    else ""
                )

                safe_label = html.escape(
                    label
                )

                bar_html += (
                    '<div '
                    'class="dist-segment" '
                    f'style="flex: {percentage} 0 auto; '
                    f'background-color: {background};" '
                    f'title="{safe_label}: '
                    f'{percentage:.2f}%">'
                    f'{safe_label}'
                    '</div>'
                )

            # -------------------------------------------------
            # OMITTED
            # -------------------------------------------------

            omitted_count = get_value(
                item,
                "omitted_count",
                0
            )

            omitted_percentage = get_value(
                item,
                "omitted_percentage",
                0
            )

            try:

                omitted_count = int(
                    omitted_count
                )

            except (
                TypeError,
                ValueError
            ):

                omitted_count = 0

            try:

                omitted_percentage = float(
                    omitted_percentage
                )

            except (
                TypeError,
                ValueError
            ):

                omitted_percentage = 0.0

            if (
                omitted_count > 0
                and
                omitted_percentage > 0
            ):

                bar_html += (
                    '<div '
                    'class="dist-segment" '
                    f'style="flex: '
                    f'{omitted_percentage} 0 auto; '
                    'background-color: #92A0BD;" '
                    f'title="Omitted: '
                    f'{omitted_percentage:.2f}%">'
                    '—'
                    '</div>'
                )

            bar_html += "</div>"

            st.markdown(
                bar_html,
                unsafe_allow_html=True
            )

            # =================================================
            # OPTION TABLE
            # =================================================

            st.markdown(
                "#### Option Breakdown"
            )

            grid_data = []

            for option in breakdown:

                is_correct = bool(
                    get_value(
                        option,
                        "is_correct",
                        False
                    )
                )

                is_functional = bool(
                    get_value(
                        option,
                        "is_functional",
                        False
                    )
                )

                if is_correct:

                    role = "✅ Correct Answer"

                elif is_functional:

                    role = "✓ Functional Distractor"

                else:

                    role = "❌ Non-functional Distractor"

                percentage = get_value(
                    option,
                    "percentage",
                    0
                )

                try:

                    percentage = float(
                        percentage
                    )

                except (
                    TypeError,
                    ValueError
                ):

                    percentage = 0.0

                grid_data.append({

                    "Option":
                        get_value(
                            option,
                            "option",
                            ""
                        ),

                    "Selection Count":
                        get_value(
                            option,
                            "count",
                            0
                        ),

                    "Distribution":
                        f"{percentage:.2f}%",

                    "Diagnostic Evaluation":
                        role
                })

            if grid_data:

                st.table(
                    grid_data
                )

            else:

                st.info(
                    "No option breakdown data was returned."
                )

    # ========================================================
    # FINAL RECOMMENDATION SUMMARY
    # ========================================================

    st.markdown("---")

    st.markdown(
        """
<div class="section-intro">
    <div class="section-kicker">Teaching actions</div>
    <div class="section-title">Item recommendation summary</div>
</div>
""",
        unsafe_allow_html=True
    )

    recommendation_rows = []

    for item in items:

        recommendation = str(
            get_value(
                item,
                "recommendation",
                ""
            )
        ).upper()

        recommendation_rows.append({

            "Question":
                get_value(
                    item,
                    "question",
                    ""
                ),

            "Difficulty":
                f"{float(get_value(item, 'difficulty', 0)):.2f}",

            "Difficulty Status":
                get_value(
                    item,
                    "difficulty_status",
                    ""
                ),

            "Discrimination":
                f"{float(get_value(item, 'discrimination', 0)):.2f}",

            "Discrimination Status":
                get_value(
                    item,
                    "discrimination_status",
                    ""
                ),

            "Distractor Efficiency":
                f"{float(get_value(item, 'distractor_efficiency', 0)):.0f}%",

            "Recommendation":
                recommendation
        })

    if recommendation_rows:

        st.dataframe(
            pd.DataFrame(
                recommendation_rows
            ),
            use_container_width=True,
            hide_index=True
        )

    # ====================================================
    # DOWNLOAD REPORT
    # ====================================================

    st.markdown("---")

    st.markdown(
        """
<div class="section-intro">
    <div class="section-kicker">Shareable record</div>
    <div class="section-title">Download the complete report</div>
</div>
<div class="download-callout">
    <strong>Your report is ready to share.</strong>
    <span>Download the Excel workbook for item-level findings, option
    breakdowns, and the overall summary with your examination information.</span>
</div>
""",
        unsafe_allow_html=True
    )

    report_bytes = create_excel_report(
        results
    )

    st.download_button(
        label="⬇️ Download Complete Item Analysis Report",
        data=report_bytes,
        file_name="Item_Analysis_Report.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        type="primary",
        use_container_width=True
    )

    st.caption(
        "The report contains Item Analysis first, followed by "
        "Option Breakdown and Overall Summary, each stamped with "
        "your exam information."
    )


# ============================================================
# UPLOAD REMINDER
# ============================================================

else:

    st.info(
        "💡 Please upload both the Answer Key and Student "
        "Responses Excel files to unlock the analysis dashboard."
    )
